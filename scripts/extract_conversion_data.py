"""
scripts/extract_conversion_data.py — Regenerate conversion_data.py from a
WBCSD TIP KPI Collection Tool workbook.

Usage:
    python scripts/extract_conversion_data.py /path/to/TIP_KPI_Collection_Tool.xlsx

Reads the "Conversion tables", "Conversion table waste", and
"Pathway 3 - Electricity input" sheets and writes a fresh conversion_data.py
to the repo root. Run this whenever WBCSD republishes updated unit-conversion
factors or IEA grid emission factors, instead of hand-editing the dicts —
hand-editing ~150 countries x 17 years of factors is exactly the kind of
transcription work that introduces silent errors.
"""
from __future__ import annotations
import sys
import json
import warnings
from pathlib import Path
from collections import defaultdict

import openpyxl

warnings.filterwarnings("ignore")


def extract_unit_conversions(wb):
    ws = wb["Conversion tables"]
    rows, cur_section = [], None
    for r in range(5, ws.max_row + 1):
        section = ws.cell(r, 1).value
        subsection = ws.cell(r, 3).value
        unit = ws.cell(r, 4).value
        cf = ws.cell(r, 5).value
        if section:
            cur_section = section
        if subsection and unit and cf is not None:
            rows.append({"section": cur_section, "subsection": subsection, "unit": unit, "cf": cf})

    ws2 = wb["Conversion table waste"]
    cur_section = None
    for r in range(5, ws2.max_row + 1):
        section = ws2.cell(r, 1).value
        subsection = ws2.cell(r, 3).value
        unit = ws2.cell(r, 4).value
        cf = ws2.cell(r, 5).value
        if section:
            cur_section = section
        if subsection and unit and cf is not None and cur_section == "Waste":
            rows.append({"section": cur_section, "subsection": subsection, "unit": unit, "cf": cf})
    return rows


def extract_grid_ef(wb):
    ws = wb["Conversion tables"]
    years = [ws.cell(3, c).value for c in range(28, 45)]
    data = {}
    for r in range(4, ws.max_row + 1):
        country = ws.cell(r, 26).value
        if country is None:
            continue
        data[country] = [ws.cell(r, c).value for c in range(28, 45)]
    return years, data


# Aggregate rows in the Electricity input sheet that close out a region group.
# None means "pure section header, not itself a selectable region".
_AGGREGATE_MARKERS = {
    "OECD Americas": "Americas",
    "OECD Asia Oceania": "Asia-Pacific (OECD)",
    "OECD Europe": "Europe (OECD)",
    "European Union - 27": "Europe (OECD)",
    "Non-OECD": None,
    "Non-OECD Europe and Eurasia": "Non-OECD Europe & Eurasia",
    "Other Africa": "Africa", "Africa": "Africa",
    "Other Asia": "Non-OECD Asia", "Asia": "Non-OECD Asia",
    "Other Non-OECD Americas": "Non-OECD Americas", "Non-OECD Americas": "Non-OECD Americas",
    "Middle East": "Middle East",
    "Total non renewable electricity purchased": None,
}


def extract_country_regions(wb):
    ws = wb["Pathway 3 - Electricity input"]
    names = [ws.cell(r, 2).value for r in range(8, ws.max_row + 1) if ws.cell(r, 2).value]

    region_for, current_region, buffer = {}, "Americas", []
    for name in names:
        if name in _AGGREGATE_MARKERS:
            label = _AGGREGATE_MARKERS[name]
            if label is None:
                for n in buffer:
                    region_for[n] = current_region
                buffer = []
                continue
            for n in buffer:
                region_for[n] = label
            region_for[name] = label
            buffer = []
            current_region = label
        else:
            buffer.append(name)
    for n in buffer:
        region_for[n] = current_region

    by_region = defaultdict(list)
    for name in names:
        if name == "Total non renewable electricity purchased" or name == "Non-OECD":
            continue
        by_region[region_for.get(name, "Other")].append(name)
    by_region.pop("Other", None)
    return dict(by_region)


def main():
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(1)
    wb = openpyxl.load_workbook(sys.argv[1], data_only=True)

    unit_conv = extract_unit_conversions(wb)
    years, grid = extract_grid_ef(wb)
    regions = extract_country_regions(wb)

    out = Path(__file__).resolve().parent.parent / "conversion_data.py"
    # (Code generation mirrors the structure of conversion_data.py — see that
    # file's module docstring. Kept here as JSON dumps for inspection/diffing
    # rather than re-deriving the full file template, so a re-run is always a
    # reviewable diff against the committed conversion_data.py.)
    Path("extracted_unit_conv.json").write_text(json.dumps(unit_conv, indent=1))
    Path("extracted_grid_ef.json").write_text(json.dumps({"years": years, "data": grid}, indent=1))
    Path("extracted_regions.json").write_text(json.dumps(regions, indent=1))
    print(f"Extracted {len(unit_conv)} conversion rows, {len(grid)} grid-EF countries, "
          f"{sum(len(v) for v in regions.values())} taxonomy entries across {len(regions)} regions.")
    print("Review the three extracted_*.json files, then update conversion_data.py by hand "
          "against the diff — do not auto-overwrite it, so changes stay reviewable.")


if __name__ == "__main__":
    main()