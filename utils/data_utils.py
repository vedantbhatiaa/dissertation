"""
utils/data_utils.py — TIP ESG Platform · Data Write Helpers
============================================================
Functions that save/load supplementary data and build master CSV rows.
Extracted from app.py to allow page modules to import without
circular dependency.

Note: _save_submission_to_csv updates in-memory globals in app.py via
state module after writing to disk.
"""
from __future__ import annotations
import pandas as pd
import numpy as np
import logging
from pathlib import Path
from datetime import datetime
from filelock import FileLock

import config as cfg
import data_loader as dl
import state

_log = logging.getLogger("esg_app")

# ── Supplementary fields now live IN the master wide CSV (retired the separate
#    ESG_SUPPLEMENTARY.csv). Each P&G/coal/water-detail field maps to a master
#    column; _load_supplementary reads them back in the legacy {supp_key: value}
#    shape so every existing caller (render_people_tab, page_entry prefill)
#    keeps working unchanged. _save_supplementary / _migrate_* are gone — the
#    save path writes these columns directly via _build_master_row(supp=...).
_SUPP_KEY_TO_MASTER_COL = {
    "stress_water_withdrawal":   "Stress Water Withdrawal",
    "non_stress_water_withdrawal": "Non-Stress Water Withdrawal",
    "coal_sub_bituminous":       "Coal Sub-Bituminous",
    "coal_brown_briquettes":     "Coal Brown Briquettes",
    "coal_other_bituminous":     "Coal Other Bituminous",
    "fuel_oil_heavy_c":          "Energy - Fuel Oil Heavy C",
    "hs_external_audit":         "HS External Audit Sites",
    "hs_internal_audit":         "HS Internal Audit Sites",
    "total_employees":           "Total Employees",
    "female_employees":          "Female Employees",
    "board_total":               "Board Total",
    "female_board":              "Female Board",
    "sbt_total":                 "SBT Total",
    "sbt_validated":             "SBT Validated",
    "sbt_committed":             "SBT Committed",
    "sbt_non_committed":         "SBT Non-Committed",
    # hs_total_sites has its own master column (added so the H&S coverage
    # denominator is the value the company actually entered, not total_sites).
    "hs_total_sites":            "HS Total Sites",
}


def _load_supplementary(company: str, year: int) -> dict:
    """Load the P&G / coal-breakdown / water-detail fields for company+year
    FROM THE MASTER WIDE CSV (supplementary file retired). Returns the legacy
    {supp_key: float} shape, or {} if the row isn't present."""
    df = state.CONSOLIDATED_DF
    if df is None or df.empty or "Company" not in df.columns or "Year" not in df.columns:
        # Fall back to reading the file directly (e.g. called before state is loaded)
        try:
            _c = next((p for p in dl._get_csv_candidates() if p.exists()
                       and p.name.startswith("ESG_MASTER_WIDE_ALL_COMPANIES_")), None)
            if not _c:
                return {}
            df = pd.read_csv(_c)
        except Exception:
            return {}
    row = df[(df["Company"] == company) & (df["Year"].astype(str) == str(year))]
    if row.empty:
        return {}
    r = row.iloc[0]
    out = {}
    for supp_key, col in _SUPP_KEY_TO_MASTER_COL.items():
        if col in df.columns:
            v = r.get(col)
            try:
                out[supp_key] = float(v) if pd.notna(v) else 0.0
            except (TypeError, ValueError):
                out[supp_key] = 0.0
    return out


def _save_supplementary(company: str, year: int, data: dict = None) -> None:
    """DEPRECATED no-op. Supplementary CSV retired; P&G fields now stored in
    master wide CSV via _save_submission_to_csv(supp=...). This shim exists
    only so legacy component imports keep working without crashing."""
    _log.debug("_save_supplementary is a deprecated no-op")
    return None


def _build_master_row(inp, out, supp: dict = None, units: dict = None,
                      corp_values: dict = None) -> dict:
    """
    Build a dict whose keys exactly match the master wide CSV column names.
    `units`: optional {field_name: unit_string} for the 23 unit-bearing
    fields in field_registry.UNIT_FIELDS. When provided, additively writes
    <WideCol>_CorporateValue and <WideCol>_Unit columns alongside the
    existing common-unit column -- nothing already reading the existing
    column names breaks, since those are untouched.
    """
    import field_registry as fr

    # M5 FIX -- was inp.coal_sub directly, which ignored the granular
    # sub-type fields and ignored fuel_oil_heavy_c outright.
    fuel_total   = (inp.nat_gas + out.total_coal + inp.propane + out.total_fuel_oil
                     + inp.diesel + inp.petrol)
    renew_share  = (inp.renew_elec_purchased + inp.self_gen_elec) / max(out.total_electricity, 1) * 100
    scope1_share = out.total_co2_scope1 / max(out.total_co2, 1) * 100
    scope2_share = out.total_co2_scope2 / max(out.total_co2, 1) * 100
    fossil_share = fuel_total / max(out.total_energy, 1) * 100
    prod         = max(inp.production, 1)

    # Waste derived
    waste_total    = float(getattr(inp, "waste_total", 0) or 0)
    waste_recovery = float(getattr(inp, "waste_recovery", 0) or 0)
    recovery_rate  = round(waste_recovery / waste_total * 100, 4) if waste_total else 0.0
    waste_elim     = round(waste_total - waste_recovery, 4)

    row = {
        "Company": inp.company, "Year": inp.year,
        "Total no. of sites": int(round(inp.total_sites)),
        "ISO 14001 sites":    int(round(inp.iso_sites)),
        "% certified sites":  round(out.pct_certified, 6),
        "Production":         round(inp.production, 4),
        "Water intake":       round(inp.water_withdrawals, 4),
        "Water intake - KPI": round(out.water_kpi, 6),
        "Total Electricity":                               round(out.total_electricity, 4),
        "Renewable Electricity Purchased":                 round(inp.renew_elec_purchased, 4),
        "Non-Renewable Electricity Purchased":             round(inp.nonrenew_elec_purchased, 4),
        "Self-generated AND consumed electricity on-site": round(inp.self_gen_elec, 4),
        "Purchased Steam":   round(inp.purchased_steam, 4),
        "Sold Electricity":  round(inp.sold_electricity, 4),
        "Sold Steam":        round(inp.sold_steam, 4),
        "Natural Gas":       round(inp.nat_gas, 4),
        "Coal":              round(out.total_coal, 4),       # M5 FIX
        "Propane":           round(inp.propane, 4),
        "Fuel Oil":          round(out.total_fuel_oil, 4),   # M5 FIX
        "Diesel":            round(inp.diesel, 4),
        "Petrol":            round(inp.petrol, 4),
        "Biomass":           round(inp.biomass, 4),
        "Waste tires":       round(inp.waste_tires_mt, 4),
        "LPG":               round(inp.lpg, 4),
        "Other":             round(inp.other_fuels, 4),
        "Energy - Sub bituminous coal":   round(inp.coal_sub_bituminous, 4),
        "Energy - Brown coal briquettes": round(inp.coal_brown_briquettes, 4),
        "Energy - Other bituminous coal": round(inp.coal_other_bituminous, 4),
        "Energy - Fuel Oil Heavy A":      round(inp.fuel_oil_heavy_a, 4),
        "Energy - Fuel Oil Heavy C":      round(inp.fuel_oil_heavy_c, 4),
        "CO2 - Sub bituminous coal":      round(out.co2_coal_sub_bituminous, 4),
        "CO2 - Brown coal briquettes":    round(out.co2_coal_brown_briquettes, 4),
        "CO2 - Other bituminous coal":   round(out.co2_coal_other_bituminous, 4),
        "CO2 - Fuel Oil Heavy A":         round(out.co2_fuel_oil_heavy_a, 4),
        "CO2 - Fuel Oil Heavy C":         round(out.co2_fuel_oil_heavy_c, 4),
        "Total energy":          round(out.total_energy, 4),
        "Total energy - KPI":    round(out.energy_kpi, 6),
        "Total CO2 - Scope 1":   round(out.total_co2_scope1, 4),
        "Total CO2 - Scope 2":   round(out.total_co2_scope2, 4),
        "Total CO2":             round(out.total_co2, 4),
        "Total CO2 - KPI":       round(out.co2_kpi, 6),
        "Total Waste":           round(waste_total, 4),
        "Waste Recovered":       round(waste_recovery, 4),
        "Recovery Rate":         recovery_rate,
        **{_elec_col(c): None for c in state.ELEC_ALL_COUNTRIES},
        "Renewable_Electricity_Share_%": round(renew_share, 4),
        "Scope1_Share_%":                round(scope1_share, 4),
        "Scope2_Share_%":                round(scope2_share, 4),
        "Fossil_Energy_Share_%":         round(fossil_share, 4),
        "Water_per_ton":                 round(inp.water_withdrawals / prod, 4),
        "CO2_per_ton":                   round(out.total_co2 / prod, 4),
        "Energy_per_ton":                round(out.total_energy / prod, 4),
        "ISO_Certification_%":           round(out.pct_certified * 100, 2),
        "Waste_Recovery_Rate_%":         recovery_rate,
        "Total_Electricity_by_Country_GJ": None,  # filled after country save
    }

    if units:
        cv = corp_values or {}
        for uf in fr.UNIT_FIELDS:
            unit = units.get(uf.field)
            if not unit:
                continue
            # Prefer the RAW value the company typed (corp_values); only fall
            # back to inp.<field> (which is already common-unit) if the raw
            # value wasn't passed through. This keeps the CorporateValue column
            # truly in the company's own unit for the corporate-units table.
            raw_value = cv.get(uf.field, getattr(inp, uf.field, None))
            row[fr.corporate_col(uf)] = round(raw_value, 4) if raw_value is not None else None
            row[fr.unit_col(uf)] = unit

    # ── People & Governance fields (from supplementary) ────────────────────
    s = supp or {}
    def _sf(k, default=0.0):
        try: return float(s.get(k, default) or default)
        except Exception: return default

    stress_wd     = _sf("stress_water_withdrawal")
    non_stress_wd = _sf("non_stress_water_withdrawal",
                         max(inp.water_withdrawals - stress_wd, 0))
    hs_ext   = _sf("hs_external_audit")
    hs_int   = _sf("hs_internal_audit")
    # H&S coverage denominator = the H&S total-sites the company entered
    # (falls back to total_sites only when not provided). Previously this
    # used hs_external_audit as the default, which made the % meaningless.
    hs_tot   = max(int(_sf("hs_total_sites", inp.total_sites)), 1)
    emp_tot  = _sf("total_employees")
    emp_fem  = _sf("female_employees")
    bod_tot  = _sf("board_total")
    bod_fem  = _sf("female_board")

    row.update({
        # Water detail
        "Stress Water Withdrawal":     round(stress_wd, 4),
        "Non-Stress Water Withdrawal": round(non_stress_wd, 4),
        # Coal breakdown
        "Coal Sub-Bituminous":         round(_sf("coal_sub_bituminous"), 4),
        "Coal Brown Briquettes":       round(_sf("coal_brown_briquettes"), 4),
        "Coal Other Bituminous":       round(_sf("coal_other_bituminous"), 4),
        # H&S
        "HS Total Sites":              round(hs_tot),
        "HS External Audit Sites":     round(hs_ext),
        "HS Internal Audit Sites":     round(hs_int),
        "HS External Audit %":         round(hs_ext / max(hs_tot, 1) * 100, 2),
        "HS Internal Audit %":         round(hs_int / max(hs_tot, 1) * 100, 2),
        # Diversity
        "Total Employees":             round(emp_tot),
        "Female Employees":            round(emp_fem),
        "Female Employees %":          round(emp_fem / max(emp_tot, 1) * 100, 2),
        "Board Total":                 round(bod_tot),
        "Female Board":                round(bod_fem),
        "Female Board %":              round(bod_fem / max(bod_tot, 1) * 100, 2),
        # Science-Based Targets
        "SBT Total":       round(_sf("sbt_total")),
        "SBT Validated":   round(_sf("sbt_validated")),
        "SBT Committed":   round(_sf("sbt_committed")),
        "SBT Non-Committed": round(_sf("sbt_non_committed")),
    })
    return row


def _save_version_parquet(inp, combined_df: pd.DataFrame) -> str:
    """
    Save the ENTIRE company template (all years) as a Parquet snapshot.
    Stored in data_storage/versions/{CompanyName}/ — subfolder only, never flat.
    Filename: CompanyName_Year_YYYYMMDD_HHMMSS.parquet (year = the year just edited).
    NEVER overwritten — each save event creates a new file (full audit trail).
    Reading this file shows the complete state of all years for that company
    at the exact moment the save was made.
    """
    from pathlib import Path
    from datetime import datetime
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    co_safe = inp.company.replace(" ", "_").replace("/", "_")
    # Extract ALL rows for this company from the combined master DataFrame
    company_all_years = combined_df[combined_df["Company"] == inp.company].copy()
    filename = f"{co_safe}_{inp.year}_{ts}.parquet"
    # Subfolder only — no flat file
    ver_dir  = Path("data_storage") / "versions" / co_safe
    ver_dir.mkdir(parents=True, exist_ok=True)
    try:
        company_all_years.to_parquet(ver_dir / filename, index=False)
        return f"{co_safe}/{filename}"
    except Exception as e:
        return f"[version save failed: {e}]"


def _drop_zero_elec_cols(df: "pd.DataFrame") -> "pd.DataFrame":
    """
    Return df with Elec_*_GJ country columns removed if every value in that
    column is zero or null across all rows.  Non-electricity columns are
    never touched.  Used so files only carry countries with actual consumption.
    """
    elec_cols = [c for c in df.columns if c.startswith("Elec_") and c.endswith("_GJ")
                 and c != "Total_Electricity_by_Country_GJ"]
    zero_cols = [c for c in elec_cols
                 if df[c].fillna(0).eq(0).all()]
    return df.drop(columns=zero_cols) if zero_cols else df


def _sync_company_member_files(master_df: "pd.DataFrame") -> list:
    """
    Write per-company CSVs in data_storage/members/TIP/<CompanyName>/<CompanyName>_latest.csv
    from the current master wide DataFrame.
    Skips any file that is locked (e.g. open in Excel) instead of crashing.
    Returns list of company names that were skipped.
    """
    from pathlib import Path
    members_tip = Path("data_storage/members/TIP")
    skipped = []
    for company, grp in master_df.groupby("Company"):
        co_safe   = str(company).replace(" ", "_")
        co_folder = members_tip / co_safe
        co_folder.mkdir(parents=True, exist_ok=True)
        try:
            # Drop electricity country columns that are all zero for this company
            grp_clean = _drop_zero_elec_cols(grp.reset_index(drop=True))
            grp_clean.to_csv(co_folder / f"{co_safe}_latest.csv", index=False)
        except (PermissionError, OSError):
            skipped.append(str(company))
    return skipped


def _update_tip_members_file(master_path: "Path", tip_master_path: "Path") -> None:
    """Rebuild the TIP members aggregate strictly from the latest master on disk.

    This prevents mismatches where the in-memory combined_df used during save
    (bootstrap/reconstruction) differs from the finally-written master CSV.
    """
    import pandas as pd
    tip_master_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        master_df = pd.read_csv(master_path)
    except Exception as e:
        _log.error("[tip_members] Could not read master to rebuild tip members: %s", e)
        return

    try:
        _drop_zero_elec_cols(master_df).to_csv(tip_master_path, index=False)
    except Exception as e:
        _log.error("[tip_members] Could not write TIP members file: %s", e)


def _migrate_supplementary_to_master() -> str:
    """
    One-time (idempotent) migration: read every row in ESG_SUPPLEMENTARY.csv
    and upsert it into the master CSV so the People & Governance columns are
    populated for all previously-submitted records.
    Safe to call on every startup — skips companies/years already promoted.
    """
    import csv as _csv
    from pathlib import Path as _P
    supp_path = _P("data_storage/master/ESG_SUPPLEMENTARY.csv")
    if not supp_path.exists():
        return "no supplementary file — skipped"

    migrated = 0
    with open(supp_path, newline="", encoding="utf-8") as f:
        rows = list(_csv.DictReader(f))

    for row in rows:
        company = row.get("Company","").strip()
        year_s  = row.get("Year","").strip()
        if not company or not year_s:
            continue
        try:
            year = int(year_s)
        except ValueError:
            continue

        # Build supp dict for this row
        supp = {k: float(v) if v else 0.0
                for k, v in row.items() if k not in ("Company","Year")}

        # Load existing TemplateInputs for this company+year
        hist = dl.get_company_hist(state.CONSOLIDATED_DF, company)
        if not hist:
            continue
        sd = dl.get_step_data(hist, year)
        sd_clean = {k: v for k, v in sd.items() if k in state.VALID_TEMPLATE_FIELDS}
        if not sd_clean:
            continue

        inp = TemplateInputs(company=company, year=year, **sd_clean)
        out = calculate(inp)

        # Re-save the master row WITH the supplementary values folded in.
        # Must pass supp= explicitly: _load_supplementary now reads the master
        # (not the legacy CSV), so an un-migrated row has empty P&G columns and
        # auto-load would drop exactly the values we're trying to migrate.
        _save_submission_to_csv(inp, out, supp=supp)
        migrated += 1

    return f"migrated {migrated} supplementary records into master"


def _write_verification_status(company: str, year: int, status: str) -> None:
    """
    Persist DSS+ verification status for a company+year to a CSV file.
    Status values: 'Verified', 'Pending', 'Flagged'.
    Client home page reads this file to show the verification chip.
    """
    from pathlib import Path
    import csv, os

    vcsv = Path("data_storage/verifications.csv")
    vcsv.parent.mkdir(parents=True, exist_ok=True)

    rows = []
    if vcsv.exists():
        with open(vcsv, newline="") as f:
            for row in csv.DictReader(f):
                if not (row.get("Company","").strip() == company and
                        str(row.get("Year","")).strip() == str(year)):
                    rows.append(row)   # keep other company/year rows

    rows.append({"Company": company, "Year": str(year), "Status": status,
                 "UpdatedBy": "dss+ Analyst"})

    with open(vcsv, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["Company","Year","Status","UpdatedBy"])
        w.writeheader()
        w.writerows(rows)


def _save_submission_to_csv(inp, out, units: dict = None, supp: dict = None,
                            corp_values: dict = None, country_values_gj: dict = None) -> str:
    """
    Three independent operations:

    1. MASTER CSV (data_storage/master/) — overwrite the row for this company+year.
       The master always holds the LATEST values. Second save for same company+year
       replaces the first row.

    2. VERSION Parquet (data_storage/versions/) — always ADD a new timestamped file.
       Never overwritten. Full audit trail of every save event.

    3. SYNC (after master is written):
       - CONSOLIDATED_DUMMY Excel Raw Dummy data sheet (long format)
       - Per-company CSVs in data_storage/members/TIP/<Company>/
       - TIP members aggregate CSV

    `units`: optional {field_name: unit_string} from page_entry.py's unit
    dropdowns — passed straight through to _build_master_row so the
    CorporateValue/Unit columns get written alongside the common-unit value.
    """
    import os, tempfile
    from pathlib import Path
    from datetime import datetime

    _master_cands = dl._get_csv_candidates()
    csv_path = next((p for p in _master_cands if p.exists()
                     and p.name.startswith("ESG_MASTER_WIDE_ALL_COMPANIES_")),
                    None) or Path("data_storage/master/ESG_MASTER_WIDE_ALL_COMPANIES_2009_2025.csv")
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    # Supplementary fields now come straight from the caller (page_entry), no
    # longer from a separate CSV. Fall back to reading them out of the master
    # row (via _load_supplementary, now master-backed) when not passed — e.g.
    # a re-save from My Records that didn't re-collect the P&G inputs.
    _supp_data = supp if supp is not None else _load_supplementary(inp.company, inp.year)
    new_row    = pd.DataFrame([_build_master_row(inp, out, supp=_supp_data,
                                                 units=units, corp_values=corp_values)])
    master_cols = list(new_row.columns)

    def _align(df):
        """Align DataFrame to master column order: strip extras, fill missing."""
        if df.empty:
            return pd.DataFrame(columns=master_cols)
        extra = [c for c in df.columns if c not in master_cols]
        if extra:
            df = df.drop(columns=extra)
        for col in master_cols:
            if col not in df.columns:
                df[col] = None
        return df[master_cols]

    def _load_best_existing():
        """
        Load the most complete existing master DataFrame.
        Checks all candidate paths and picks the one with the most rows.
        """
        candidates = [
            csv_path,
            Path("data_storage/raw/ESG_MASTER_WIDE_ALL_COMPANIES_2009_2025.csv"),
        ]
        best = pd.DataFrame(columns=master_cols)
        for p in candidates:
            if p.exists():
                try:
                    df = pd.read_csv(p)
                    if "Company" in df.columns and "Year" in df.columns and len(df) > len(best):
                        best = df
                except PermissionError:
                    pass
                except Exception:
                    pass
        return _align(best)

    # ── 1. Build combined DataFrame ──────────────────────────────────────────
    # H1 FIX: advisory file lock held for the full read-modify-write cycle.
    # Prevents data corruption if two analysts save simultaneously.
    # Timeout=10s: if another process holds the lock and crashes, we don't
    # block forever.  The PermissionError branch below still handles Excel locks.
    lock_path = csv_path.with_suffix(".lock")
    with FileLock(str(lock_path), timeout=cfg.FILELOCK_TIMEOUT):
        existing = _load_best_existing()
        mask     = ~((existing["Company"] == inp.company) & (existing["Year"] == inp.year))
        existing = existing[mask]
        combined = pd.concat([existing, new_row], ignore_index=True)
        combined = combined.sort_values(["Company", "Year"]).reset_index(drop=True)

        n_records   = len(combined)
        n_companies = combined["Company"].nunique()

        # ── 2. Save version Parquet BEFORE touching master (audit trail first) ───
        version_filename = _save_version_parquet(inp, combined)

        # ── 3. Write master CSV, then sync all dependent files ───────────────────
        try:
            # Master CSV keeps all country columns (even all-zero) as the full schema.
            # Derived outputs (member files, TIP aggregate) strip all-zero country cols.
            combined.to_csv(csv_path, index=False)
            # Sync TIP members aggregate
            tip_master_path = Path("data_storage/members/TIP/ESG_MASTER_WIDE_TIP_MEMBERS_2009_2023.csv")
            _update_tip_members_file(csv_path, tip_master_path)
            # Sync per-company member files
            _sync_company_member_files(combined)
            # Sync CONSOLIDATED_DUMMY Excel
            _sync_consolidate_excel(combined)

            # ── Auto-add electricity-by-country year columns for new submission ──
            # If company just submitted for a year that isn't in the electricity
            # editor yet, initialize all country columns to 0 in the master CSV.
            _new_yr = inp.year
            _elec_cols_all = [c for c in combined.columns
                              if c.startswith("Elec_") and c.endswith("_GJ")]
            _co_yr_mask = (combined["Company"] == inp.company) & (combined["Year"] == _new_yr)
            if _co_yr_mask.any() and _elec_cols_all:
                for _ec in _elec_cols_all:
                    if pd.isna(combined.loc[_co_yr_mask, _ec]).all():
                        combined.loc[_co_yr_mask, _ec] = 0.0
                # Also ensure Total_Electricity_by_Country_GJ exists
                if "Total_Electricity_by_Country_GJ" not in combined.columns:
                    combined["Total_Electricity_by_Country_GJ"] = 0.0
                elif pd.isna(combined.loc[_co_yr_mask, "Total_Electricity_by_Country_GJ"]).all():
                    combined.loc[_co_yr_mask, "Total_Electricity_by_Country_GJ"] = 0.0
                # Re-write master with the zero-initialized electricity columns
                combined.to_csv(csv_path, index=False)

            # ── 3b. Update the LONG-FORM master for this company+year ───────────
            # Wide and long are kept in lock-step: every Submit & Save overwrites
            # this company+year in both files. Long form is the canonical 224-
            # field-per-year schema (Value in corporate unit, GroundTruthCommon-
            # Value in common unit). Failure here must not abort the wide save,
            # but it must be VISIBLE — silently logging it where no one reads
            # server logs is exactly how wide/long drifted out of sync before.
            try:
                _lf_status = _upsert_longform(inp, out, units=units, corp_values=corp_values,
                                              country_values_gj=country_values_gj, supp=_supp_data)
            except Exception as _lf_e:
                _lf_status = f"long-form update FAILED: {_lf_e}"
                _log.warning("[save] %s", _lf_status)

            # ── CRITICAL: update the in-memory globals so all pages in this
            #    session immediately see the new data without requiring a restart.
            state.CONSOLIDATED_DF     = combined.copy()
            state.COMPANIES           = dl.get_companies(combined)
            state.USING_FALLBACK = False
            try:
                cfg.refresh_year_bounds(combined)
                state.HIST_YEARS = cfg.hist_years()
                state.CURR_YEAR  = cfg.curr_year()
                state.LONG_YEARS = cfg.long_years()
            except Exception:
                pass
            try:
                state.SECTOR_DF = dl.load_sector_aggregated(combined)
            except Exception:
                pass
            # LONG_DATA/FUEL_MIX are rebuilt by app.py after state update
            try:
                import streamlit as _st
                _st.cache_data.clear()
                # Bust get_hist_outputs session cache so pages re-compute immediately.
                # Delete by stored key name — avoids iterating all session_state keys.
                _cached_key = _st.session_state.get("_hist_cache_last_key")
                if _cached_key:
                    _st.session_state.pop(_cached_key, None)
                    _st.session_state.pop("_hist_cache_last_key", None)
            except Exception: pass

            return (f"Saved {inp.company} — {inp.year}. "
                    f"Master: {n_records} records across {n_companies} companies. "
                    f"Version: {version_filename}. {_lf_status.capitalize()}.")
        except PermissionError:
            ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"ESG_MASTER_{inp.company.replace(' ','_')}_{inp.year}_{ts}.csv"
            backup_path = csv_path.parent / backup_name
            try:
                combined.to_csv(backup_path, index=False)
                return (
                    f"⚠️ Master file open in Excel — saved backup: **{backup_name}**\n"
                    f"Version snapshot: {version_filename}\n"
                    f"Close Excel and click Save again."
                )
            except Exception as e2:
                return f"❌ Save failed (file locked AND backup failed): {e2}"
        except Exception as e:
            return f"❌ Save failed: {e}"


def _add_iea_co2_columns(master: "pd.DataFrame", mask, year: int, country_col_map: dict) -> None:
    """Compute and write IEA_CO2_<Country>_tCO2 columns for the row(s)
    selected by `mask`, from the Elec_<Country>_GJ columns that were just
    written and the IEA location-based grid emission factors already in
    conversion_data.ELECTRICITY_GRID_EF.

    tCO2 = MWh × EF(g CO2/kWh) / 1000   (EF tables are in g/kWh; MWh×1000=kWh,
    and grams → tonnes is ÷1,000,000, so the two conversions net to ÷1000.)

    Column names mirror render_electricity_tab.py's parsing exactly — it
    recovers the country name via
    `c.replace("IEA_CO2_","").replace("_tCO2","").replace("_"," ")`,
    so the column must be built by swapping the Elec_/_GJ affixes on the
    SAME column name string, not by re-deriving from the country name
    (which could drift on punctuation like "Côte d'Ivoire").
    """
    import conversion_data as cd
    if not mask.any():
        return
    for country, gj_col in country_col_map.items():
        if gj_col not in master.columns or not gj_col.startswith("Elec_") or not gj_col.endswith("_GJ"):
            continue
        ef = cd.get_grid_ef(country, year)
        if ef is None:
            continue  # no IEA factor for this country/region — leave column absent
        iea_col = "IEA_CO2_" + gj_col[len("Elec_"):-len("_GJ")] + "_tCO2"
        if iea_col not in master.columns:
            master[iea_col] = None
        gj_val  = pd.to_numeric(master.loc[mask, gj_col], errors="coerce").fillna(0.0)
        mwh_val = gj_val / 3.6
        master.loc[mask, iea_col] = (mwh_val * ef / 1000.0).round(4)


def _elec_col_for_country(country: str) -> str | None:
    """Resolve the Elec_<Country>_GJ column for ANY country in the full IEA
    taxonomy (state.ELEC_ALL_COUNTRIES) — not just the subset state.
    ELEC_COUNTRY_COLS happens to contain. That dict's own inline comments
    say "all 31 countries"; it appears to predate the ~150-country taxonomy
    the editor itself now displays, so any country outside the original 31
    (Israel being the confirmed case) was being silently skipped on every
    save (`continue`) even though the UI happily accepts input for it.
    Falls back to the same _elec_col() naming convention already used for
    every existing Elec_* column, so this is a strict superset of the old
    behaviour — anything that worked before still resolves the same way.
    Returns None only if `country` isn't in the recognised taxonomy at all
    (guards against typos / garbage input being turned into new columns).
    """
    col = state.ELEC_COUNTRY_COLS.get(country)
    if col:
        return col
    if country in getattr(state, "ELEC_ALL_COUNTRIES", []):
        return _elec_col(country)
    return None


def _save_electricity_to_master(company: str, year: int) -> str:
    """
    Save electricity-by-country data (from the Electricity tab editor) into:
      1. Master wide CSV  — columns Elec_<Country>_GJ  (GJ = MWh x 3.6)
      2. TIP members aggregate CSV
      3. Per-company member CSVs in data_storage/members/TIP/<Company>/
      4. CONSOLIDATED_DUMMY Excel (Raw Dummy data sheet, long format)
      5. Long-form master — "Corporate units - <Country>" KPI rows
      6. Parquet snapshot of the complete company+year row

    Updates ALL years present in the electricity editor. Every country in
    state.ELEC_ALL_COUNTRIES (~150, the full IEA taxonomy) is written, via
    _elec_col_for_country() — columns are auto-created in the master CSV
    if they don't exist yet.
    """
    import streamlit as st
    from pathlib import Path
    from datetime import datetime

    ALL_COUNTRIES = state.ELEC_ALL_COUNTRIES
    MWH_TO_GJ = 3.6

    elec_df = st.session_state.get("elec_data", pd.DataFrame())
    if elec_df.empty:
        return "No electricity data entered yet."

    _ecands = dl._get_csv_candidates()
    csv_path = next((p for p in _ecands if p.exists()
                     and p.name.startswith("ESG_MASTER_WIDE_ALL_COMPANIES_")), None)
    if csv_path is None:
        return "Master CSV not found. Save KPI data first."
    try:
        master = pd.read_csv(csv_path)
    except PermissionError:
        return "Master CSV is open in Excel — close it and try again."

    # Resolve every country to its column ONCE — covers the full taxonomy,
    # not just the legacy ~31-country subset (see _elec_col_for_country).
    country_to_col = {c: _elec_col_for_country(c) for c in ALL_COUNTRIES}
    country_to_col = {c: col for c, col in country_to_col.items() if col}
    all_cols = list(country_to_col.values())

    # Ensure all country columns exist in master (add if missing)
    for col in all_cols:
        if col not in master.columns:
            master[col] = None
    if "Total_Electricity_by_Country_GJ" not in master.columns:
        master["Total_Electricity_by_Country_GJ"] = None

    yr_cols = [c for c in elec_df.columns if str(c).isdigit() and 2000 < int(c) < 2030]

    updated_years, _lf_failures = [], []
    for yr_str in yr_cols:
        yr   = int(yr_str)
        mask = (master["Company"] == company) & (master["Year"] == yr)
        if not mask.any():
            # KPI row not found for this year — create a minimal stub row so
            # electricity data is not lost; user can submit KPIs later.
            stub = pd.DataFrame([{
                "Company": company, "Year": yr,
                **{c: 0.0 for c in all_cols},
                "Total_Electricity_by_Country_GJ": 0.0,
            }])
            # Align to master columns
            for col in master.columns:
                if col not in stub.columns:
                    stub[col] = None
            stub = stub[master.columns]
            master = pd.concat([master, stub], ignore_index=True)
            master = master.sort_values(["Company", "Year"]).reset_index(drop=True)
            mask = (master["Company"] == company) & (master["Year"] == yr)

        year_series = elec_df.set_index("Country")[yr_str]
        country_vals_gj = {}
        for country, mwh_val in year_series.items():
            col_name = country_to_col.get(str(country))
            if col_name is None:
                continue  # not in the recognised taxonomy at all
            gj_val = float(mwh_val) * MWH_TO_GJ if pd.notna(mwh_val) else 0.0
            master.loc[mask, col_name] = round(gj_val, 4)
            country_vals_gj[str(country)] = round(gj_val, 4)

        # Recompute total-by-country for this row
        country_vals = [master.loc[mask, c].values[0]
                        for c in all_cols if c in master.columns]
        master.loc[mask, "Total_Electricity_by_Country_GJ"] = round(
            sum(v for v in country_vals if pd.notna(v)), 4)

        # IEA CO2 — derived straight from the GJ columns just written above,
        # using the grid emission factors in conversion_data.py. This is what
        # the "CO2 Emissions from IEA" tab reads; without this it stayed
        # permanently empty no matter what electricity data was entered.
        _add_iea_co2_columns(master, mask, yr, country_to_col)

        # Long-form sync — this save path previously never touched the
        # long-form master at all (only _save_submission_to_csv did), which
        # is why wide/consolidate/parquet updated but long-form didn't.
        # Best-effort: reconstruct (inp, out) for this company+year from the
        # master row via _load_company_year_outputs (already used elsewhere
        # in this codebase for historical display) and upsert just the
        # "Corporate units - <Country>" KPI rows. If that helper's signature
        # doesn't match what's assumed here, this is caught and logged
        # rather than failing the wide-CSV save that already succeeded.
        try:
            from utils.helpers import _load_company_year_outputs
            _inp_lf, _out_lf = _load_company_year_outputs(company, yr)
            _upsert_longform(_inp_lf, _out_lf, country_values_gj=country_vals_gj)
        except Exception as _lf_e:
            _lf_failures.append(yr)
            _log.warning("[_save_electricity_to_master] long-form sync failed for %s %s: %s",
                        company, yr, _lf_e)

        updated_years.append(yr)

    if not updated_years:
        return f"No KPI rows found for {company}. Save KPI data first."

    try:
        # H1 FIX: use the same advisory lock as the KPI save path
        lock_path = csv_path.with_suffix(".lock")
        with FileLock(str(lock_path), timeout=cfg.FILELOCK_TIMEOUT):
            master.to_csv(csv_path, index=False)
    except PermissionError:
        return "Master CSV is open in Excel — close it and try again."

    # Sync all dependent files
    try:
        _sp = csv_path.stem.split("_"); _ys, _ye = _sp[-2], _sp[-1]
    except Exception:
        _ys, _ye = str(cfg.DATA_YEAR_START), str(cfg.DATA_YEAR_END)
    tip_master_path = Path(f"data_storage/members/TIP/ESG_MASTER_WIDE_TIP_MEMBERS_{_ys}_{_ye}.csv")
    _update_tip_members_file(csv_path, tip_master_path)
    _sync_company_member_files(master)
    _sync_consolidate_excel(master)

    # Parquet snapshot
    co_safe  = company.replace(" ", "_").replace("/", "_")
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    ver_dir  = Path("data_storage") / "versions" / co_safe
    ver_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{co_safe}_{year}_elec_{ts}.parquet"
    single_row = master[(master["Company"] == company) & (master["Year"] == year)].copy()
    try:
        single_row.to_parquet(ver_dir / filename, index=False)
    except Exception:
        filename = "[parquet skipped]"

    # CRITICAL: refresh the in-memory state — this was missing entirely
    # before, which is why a newly-added country (or any edit made here)
    # never showed up in Common Units / IEA CO2 within the same session.
    # Those tabs read straight from state.CONSOLIDATED_DF, not from disk.
    state.CONSOLIDATED_DF = master.copy()
    state.COMPANIES       = dl.get_companies(master)
    try:
        import streamlit as _st
        _st.cache_data.clear()
        _cached_key = _st.session_state.get("_hist_cache_last_key")
        if _cached_key:
            _st.session_state.pop(_cached_key, None)
            _st.session_state.pop("_hist_cache_last_key", None)
        _st.session_state.pop("_elec_load_key", None)
    except Exception:
        pass

    _lf_msg = ("long-form synced" if not _lf_failures
               else f"long-form sync FAILED for {len(_lf_failures)} year(s): {_lf_failures}")
    return (f"Electricity saved — {len(updated_years)} year(s) updated "
            f"({min(updated_years)}-{max(updated_years)}) converted MWh to GJ. "
            f"Consolidate + member files synced. {_lf_msg}. "
            f"Snapshot: versions/{co_safe}/{filename}")


def _save_country_electricity_values(company: str, year: int, values_gj: dict) -> str:
    """
    Write ONE year's worth of per-country electricity values (GJ) directly
    into the master CSV row for company+year.

    Used by the Submit Data form's "Electricity by Country" section, where the
    person is editing a single reporting year inline (unlike the dedicated
    Electricity by Country tab's multi-year grid, which calls
    _save_electricity_to_master instead). Call this AFTER _save_submission_to_csv
    so the company+year row already exists (that function auto-zero-fills all
    Elec_*_GJ columns for a brand-new row before this overwrites them).

    values_gj: {country_name: value_in_GJ}
    """
    ALL_COUNTRIES  = state.ELEC_ALL_COUNTRIES
    country_to_col = {c: _elec_col_for_country(c) for c in ALL_COUNTRIES}
    country_to_col = {c: col for c, col in country_to_col.items() if col}
    all_cols       = list(country_to_col.values())

    _ecands = dl._get_csv_candidates()
    csv_path = next((p for p in _ecands if p.exists()
                     and p.name.startswith("ESG_MASTER_WIDE_ALL_COMPANIES_")), None)
    if csv_path is None:
        return "Master CSV not found — save main KPI data first."
    try:
        master = pd.read_csv(csv_path)
    except PermissionError:
        return "Master CSV is open in Excel — close it and try again."

    for col in all_cols:
        if col not in master.columns:
            master[col] = None
    if "Total_Electricity_by_Country_GJ" not in master.columns:
        master["Total_Electricity_by_Country_GJ"] = None

    mask = (master["Company"] == company) & (master["Year"] == year)
    if not mask.any():
        return f"No KPI row found for {company} — {year}. Save main KPI data first."

    for country, gj_val in values_gj.items():
        col_name = country_to_col.get(country) or _elec_col_for_country(str(country))
        if col_name is None:
            continue  # not a recognised country — ignore rather than fail the whole save
        if col_name not in master.columns:
            master[col_name] = None
        master.loc[mask, col_name] = round(float(gj_val or 0), 4)

    country_vals = [master.loc[mask, c].values[0]
                     for c in all_cols if c in master.columns]
    master.loc[mask, "Total_Electricity_by_Country_GJ"] = round(
        sum(v for v in country_vals if pd.notna(v)), 4)

    # IEA CO2 — see _add_iea_co2_columns docstring. Without this the
    # "CO2 Emissions from IEA" tab stays empty regardless of which of the
    # two electricity-save paths the company used.
    _add_iea_co2_columns(master, mask, year, country_to_col)

    try:
        lock_path = csv_path.with_suffix(".lock")
        with FileLock(str(lock_path), timeout=cfg.FILELOCK_TIMEOUT):
            master.to_csv(csv_path, index=False)
    except PermissionError:
        return "Master CSV is open in Excel — close it and try again."
    except Exception as e:
        _log.warning("[_save_country_electricity_values] %s", e)
        return f"Electricity by country save failed: {e}"

    # Sync dependents + refresh in-memory state immediately, same as the
    # other save paths — so My Records / Company Data / the Electricity tab
    # all see the update this session without needing a restart.
    try:
        _sp = csv_path.stem.split("_"); _ys, _ye = _sp[-2], _sp[-1]
    except Exception:
        _ys, _ye = str(cfg.DATA_YEAR_START), str(cfg.DATA_YEAR_END)
    tip_master_path = Path(f"data_storage/members/TIP/ESG_MASTER_WIDE_TIP_MEMBERS_{_ys}_{_ye}.csv")
    try:
        _update_tip_members_file(csv_path, tip_master_path)
        _sync_company_member_files(master)
        _sync_consolidate_excel(master)
    except Exception as e:
        _log.warning("[_save_country_electricity_values] sync: %s", e)

    state.CONSOLIDATED_DF = master.copy()
    try:
        import streamlit as _st
        _st.cache_data.clear()
        # Bust get_hist_outputs cache + force electricity editor reload
        _cached_key = _st.session_state.get("_hist_cache_last_key")
        if _cached_key:
            _st.session_state.pop(_cached_key, None)
            _st.session_state.pop("_hist_cache_last_key", None)
        _st.session_state.pop("_elec_load_key", None)
    except Exception:
        pass

    return f"Electricity by country saved for {company} — {year}."



def seed_elec_session_state(company: str, year: int) -> None:
    """Seed st.session_state["elec_data"] from the master CSV for company+year.
    Call this at the TOP of render_electricity_tab() (before the data_editor)
    so the grid always shows master data, not a stale empty DataFrame.
    Keyed by (company, year) so switching company/year refreshes automatically.

    elec_data schema: DataFrame with index=Country, columns=[year_str].
    The tab editor stores MWh (display unit); master stores GJ.
    We convert GJ → MWh here so the editor stays in MWh by default.
    """
    import streamlit as _st
    GJ_TO_MWH = 1 / 3.6
    cache_key  = f"_elec_seeded_{company}_{year}"
    if _st.session_state.get(cache_key):
        return   # already seeded this company+year this session

    COUNTRY_COL = state.ELEC_COUNTRY_COLS   # {country_name: "Elec_<Country>_GJ"}
    df = state.CONSOLIDATED_DF
    if df is None or df.empty or "Company" not in df.columns:
        return

    row = df[(df["Company"] == company) & (df["Year"].astype(str) == str(year))]
    if row.empty:
        return

    r = row.iloc[0]
    data = {}
    for country, col in COUNTRY_COL.items():
        if col in df.columns:
            v = r.get(col)
            try:
                data[country] = round(float(v) * GJ_TO_MWH, 4) if pd.notna(v) else 0.0
            except (TypeError, ValueError):
                data[country] = 0.0
        else:
            data[country] = 0.0

    # Build elec_data DataFrame: one row per country, one column per year.
    # If an existing elec_data has other years, preserve them.
    yr_str = str(year)
    existing = _st.session_state.get("elec_data", pd.DataFrame())
    if not existing.empty and "Country" in existing.columns:
        elec_df = existing.set_index("Country") if "Country" in existing.columns else existing.copy()
    elif not existing.empty:
        elec_df = existing.copy()
    else:
        elec_df = pd.DataFrame(index=list(COUNTRY_COL.keys()))

    elec_df.index.name = "Country"
    elec_df[yr_str] = pd.Series(data)
    elec_df = elec_df.reset_index()
    _st.session_state["elec_data"] = elec_df
    _st.session_state[cache_key]   = True


def _sync_consolidate_excel(master_df: "pd.DataFrame") -> None:
    """
    Sync the CONSOLIDATED_DUMMY Excel (Raw Dummy data sheet) from the master wide CSV.

    The Raw Dummy data sheet stores data in long format: one row per
    (Company, Year, Row_Label). This function overwrites it completely from
    the current master wide DataFrame so that the consolidate stays in sync
    after any save from the platform.
    """
    from pathlib import Path
    from openpyxl import load_workbook

    xl_path = Path("data_storage/master/CONSOLIDATED_DUMMY_2009_2023.xlsx")
    if not xl_path.exists():
        return  # nothing to sync yet

    # Mapping: wide-CSV column  →  (Section, Row_Label)
    COL_MAP = {
        "Total no. of sites":                              ("ISO 14001",    "Total no. of sites"),
        "ISO 14001 sites":                                 ("ISO 14001",    "ISO 14001 sites"),
        "% certified sites":                               ("ISO 14001",    "% certified sites"),
        "Production":                                      ("Production",   "Production"),
        "Water intake":                                    ("Water",        "Water intake"),
        "Water intake - KPI":                              ("Water",        "Water intake - KPI"),
        "Total Electricity":                               ("Energy",       "Total Electricity"),
        "Renewable Electricity Purchased":                 ("Energy",       "Renewable Electricity Purchased"),
        "Non-Renewable Electricity Purchased":             ("Energy",       "Non-Renewable Electricity Purchased"),
        "Self-generated AND consumed electricity on-site": ("Energy",       "Self-generated AND consumed electricity on-site"),
        "Purchased Steam":                                 ("Energy",       "Purchased Steam"),
        "Sold Electricity":                                ("Energy",       "Sold Electricity"),
        "Sold Steam":                                      ("Energy",       "Sold Steam"),
        "Natural Gas":                                     ("Energy",       "Natural Gas"),
        "Coal":                                            ("Energy",       "Coal"),
        "Propane":                                         ("Energy",       "Propane"),
        "Fuel Oil":                                        ("Energy",       "Fuel Oil"),
        "Diesel":                                          ("Energy",       "Diesel"),
        "Petrol":                                          ("Energy",       "Petrol"),
        "Biomass":                                         ("Energy",       "Biomass"),
        "Waste tires":                                     ("Energy",       "Waste tires"),
        "LPG":                                             ("Energy",       "LPG"),
        "Other":                                           ("Energy",       "Other"),
        "Total energy":                                    ("Energy",       "Total energy"),
        "Total energy - KPI":                              ("Energy",       "Total energy - KPI"),
        "Total CO2 - Scope 1":                             ("CO2 emissions","Total CO2 - Scope 1"),
        "Total CO2 - Scope 2":                             ("CO2 emissions","Total CO2 - Scope 2"),
        "Total CO2":                                       ("CO2 emissions","Total CO2"),
        "Total CO2 - KPI":                                 ("CO2 emissions","Total CO2 - KPI"),
        "Total Waste":                                     ("Waste",        "Total Waste"),
        "Waste Recovered":                                 ("Waste",        "Waste Recovered"),
        "Recovery Rate":                                   ("Waste",        "Recovery Rate"),
        **{_elec_col(c): ("Energy", f"Electricity - {c}") for c in state.ELEC_ALL_COUNTRIES},
        # People & Governance (promoted from supplementary)
        "Stress Water Withdrawal":     ("Water",       "Stress water withdrawal"),
        "Non-Stress Water Withdrawal": ("Water",       "Non-stress water withdrawal"),
        "Coal Sub-Bituminous":         ("Energy",      "Coal — Sub-bituminous"),
        "Coal Brown Briquettes":       ("Energy",      "Coal — Brown briquettes"),
        "Coal Other Bituminous":       ("Energy",      "Coal — Other bituminous"),
        "HS External Audit Sites":     ("H&S",         "Externally audited H&S sites"),
        "HS Internal Audit Sites":     ("H&S",         "Internally audited H&S sites"),
        "HS External Audit %":         ("H&S",         "H&S external audit coverage %"),
        "HS Internal Audit %":         ("H&S",         "H&S internal audit coverage %"),
        "Total Employees":             ("Diversity",   "Total employees"),
        "Female Employees":            ("Diversity",   "Female employees"),
        "Female Employees %":          ("Diversity",   "% Female employees"),
        "Board Total":                 ("Diversity",   "Board of Directors total"),
        "Female Board":                ("Diversity",   "Female Board members"),
        "Female Board %":              ("Diversity",   "% Female Board"),
        "SBT Total":                   ("SBT",         "Total with science-based target"),
        "SBT Validated":               ("SBT",         "SBT — Validated"),
        "SBT Committed":               ("SBT",         "SBT — Committed"),
        "SBT Non-Committed":           ("SBT",         "SBT — Non-committed"),
    }

    # Build long rows from master_df
    long_rows = []  # list of dicts: Company, Row, Year, Data, Section, Row_Label, Notes, Consistency test
    row_order = list(COL_MAP.keys())
    # Assign fixed row numbers to match what build_esg_master.py uses
    ROW_NUM = {col: i + 1 for i, col in enumerate(row_order)}

    # Pre-compute which electricity country columns have any non-zero value
    # across the whole master — only those countries get rows in the consolidate.
    active_elec_cols = {
        col for col in COL_MAP
        if col.startswith("Elec_") and col.endswith("_GJ")
        and col in master_df.columns
        and master_df[col].fillna(0).ne(0).any()
    }

    for _, wrow in master_df.sort_values(["Company", "Year"]).iterrows():
        company = wrow["Company"]
        year    = int(wrow["Year"]) if pd.notna(wrow.get("Year")) else None
        if not company or not year:
            continue
        for col, (section, label) in COL_MAP.items():
            # Skip electricity country columns that are all-zero across the dataset
            is_elec_country = col.startswith("Elec_") and col.endswith("_GJ")
            if is_elec_country and col not in active_elec_cols:
                continue
            val = wrow.get(col)
            # For an active electricity country, skip rows where this company-year is zero
            if is_elec_country and (pd.isna(val) or float(val) == 0):
                continue
            long_rows.append({
                "Company": company,
                "Row":     ROW_NUM[col],
                "Year":    year,
                "Data":    float(val) if pd.notna(val) else None,
                "Section": section,
                "Row_Label": label,
                "Notes":   None,
                "Consistency test": None,
            })

    if not long_rows:
        return

    try:
        wb = load_workbook(xl_path)
        ws = wb["Raw Dummy data"]
        # Clear existing data rows (keep header row 1)
        for r in range(2, ws.max_row + 1):
            for c in range(1, 9):
                ws.cell(r, c).value = None
        # Write new rows
        cols = ["Company", "Row", "Year", "Data", "Section", "Row_Label", "Notes", "Consistency test"]
        for i, row in enumerate(long_rows):
            for j, col in enumerate(cols):
                ws.cell(i + 2, j + 1).value = row[col]
        wb.save(xl_path)
    except (PermissionError, OSError):
        pass  # file locked — skip, master CSV is the source of truth
    except Exception:
        pass  # any other error is also non-fatal


def _elec_col(country: str) -> str:
    """Canonical master CSV column name for a country's electricity (GJ)."""
    return "Elec_" + country.replace(" ", "_") + "_GJ"

# ═══════════════════════════════════════════════════════════════════════════
# LONG-FORM MASTER (canonical 224-field-per-company-year schema)
# ═══════════════════════════════════════════════════════════════════════════
# Wide and long masters are kept in lock-step by _save_submission_to_csv:
# every Submit & Save overwrites the submitted company+year in BOTH files.
#   long form columns: Company, Year, KPI_Name, Value, Unit, GroundTruthCommonValue
#     Value                  = the number in the company's reporting unit
#     Unit                   = that reporting unit (only the 23 unit-bearing KPIs)
#     GroundTruthCommonValue = the value converted to the common unit
LONG_FORM_PATH = Path("data_storage/master/ESG_LONG_FORM_MASTER.csv")
LONG_FORM_COLS = ["Company", "Year", "KPI_Name", "Value", "Unit", "GroundTruthCommonValue"]


def _canonical_kpi_order():
    """Canonical KPI_Name order, read from the existing long-form master (the
    schema source of truth). Returns [] if the file can't be read — callers
    then skip the long-form write rather than guessing an order."""
    try:
        p = LONG_FORM_PATH
        if not p.exists():
            cands = [c for c in dl._get_csv_candidates()
                     if c.name.startswith("ESG_LONG_FORM_MASTER")]
            p = next((c for c in cands if c.exists()), None)
        if not p or not p.exists():
            return []
        df = pd.read_csv(p, usecols=["Company", "Year", "KPI_Name"])
        first = df[(df["Company"] == df["Company"].iloc[0]) & (df["Year"] == df["Year"].iloc[0])]
        return list(first["KPI_Name"])
    except Exception:
        return []


def build_longform_rows(order, inp, out, units=None, corp_values=None,
                        country_values_gj=None, supp=None):
    """Build the long-form rows (one per KPI_Name in `order`) for one
    company+year submission. Mirrors _build_master_row's value sources so
    wide and long never disagree."""
    import field_registry as fr
    import conversion_data as cd
    units = units or {}; corp_values = corp_values or {}
    cvals = country_values_gj or {}; s = supp or {}

    def sf(k, d=0.0):
        try: return float(s.get(k, d) or d)
        except Exception: return d

    wt   = float(getattr(inp, "waste_total", 0) or 0)
    prod = max(inp.production, 1)
    hs_ext = sf("hs_external_audit"); hs_int = sf("hs_internal_audit")
    hs_tot = max(sf("hs_total_sites", inp.total_sites), 1.0)
    emp = sf("total_employees"); femp = sf("female_employees")
    bod = sf("board_total"); fbod = sf("female_board")

    VALUE = {
        "Total no. of sites": inp.total_sites, "ISO 14001 sites": inp.iso_sites,
        "% certified sites": out.pct_certified, "Water intake - KPI": out.water_kpi,
        "Total Purchased Electricity": out.total_electricity,
        "Energy - Total Coal": out.total_coal, "Energy - Fuel Oil": out.total_fuel_oil,
        "Total energy": out.total_energy, "Total energy - KPI": out.energy_kpi,
        "CO2 - Natural Gas": out.co2_nat_gas,
        "CO2 - Sub bituminous coal": out.co2_coal_sub_bituminous,
        "CO2 - Brown coal briquettes": out.co2_coal_brown_briquettes,
        "CO2 - Other bituminous coal": out.co2_coal_other_bituminous,
        "CO2 - Coal": out.co2_coal, "CO2 - Propane": out.co2_propane,
        "CO2 - Fuel Oil Heavy A": out.co2_fuel_oil_heavy_a,
        "CO2 - Fuel Oil Heavy C": out.co2_fuel_oil_heavy_c,
        "CO2 - Fuel Oil": out.co2_fuel_oil, "CO2 - Diesel": out.co2_diesel,
        "CO2 - Petrol": out.co2_petrol, "CO2 - Biomass": out.co2_biomass,
        "CO2 - Waste Tires": out.co2_waste_tires, "CO2 - LPG": out.co2_lpg,
        "CO2 - Other": out.co2_other,
        "Total CO2 - Scope 1": out.total_co2_scope1,
        "Total CO2 - Scope 2": out.total_co2_scope2,
        "Total CO2": out.total_co2, "Total CO2 - KPI": out.co2_kpi,
        "Waste Production": inp.production,
        "Amount of waste sent to elimination": out.waste_elimination,
        "Waste recovery %": out.waste_recovery_pct,
        "Waste elimination %": (1 - out.waste_recovery_pct) if wt else 0.0,
        "Waste Intensity - KPI": (wt / prod),
        "HS External Audit Sites": hs_ext, "HS Internal Audit Sites": hs_int,
        "HS External Audit %": round(hs_ext / hs_tot * 100, 2),
        "HS Internal Audit %": round(hs_int / hs_tot * 100, 2),
        "Total Employees": emp, "Female Employees": femp,
        "Female Employees %": round(femp / max(emp, 1) * 100, 2),
        "Board Total": bod, "Female Board": fbod,
        "Female Board %": round(fbod / max(bod, 1) * 100, 2),
        "SBT Total": sf("sbt_total"), "SBT Validated": sf("sbt_validated"),
        "SBT Committed": sf("sbt_committed"), "SBT Non-Committed": sf("sbt_non_committed"),
    }
    cu_total = sum(v for v in cvals.values() if v)

    rows, failed_kpis = [], []
    for kpi in order:
        value = unit = ground = None
        try:
            uf = fr.BY_KPI_NAME.get(kpi)
            if uf is not None:                                  # 23 unit-bearing KPIs
                common = float(getattr(inp, uf.field, 0) or 0)
                unit = units.get(uf.field)
                if uf.field in corp_values:
                    value = corp_values[uf.field]
                elif unit:
                    cf = cd.get_unit_conversion_factor(uf.subsection, unit)
                    value = common / cf if cf else common
                else:
                    value = common
                ground = common
            elif kpi.startswith("Corporate units -"):
                if kpi.endswith("Total non renewable electricity purchased"):
                    value = cu_total
                else:
                    cname = kpi.replace("Corporate units - ", "").strip().lower()
                    match = next((c for c in cvals if str(c).strip().lower() == cname), None)
                    value = cvals.get(match) if match else None
            else:
                value = VALUE.get(kpi)
        except Exception as _kpi_e:
            # Don't let one bad KPI (e.g. an unrecognized unit/subsection
            # combination) silently kill the whole 224-row block — every
            # other KPI for this company+year still gets written correctly,
            # and the caller can report exactly which one(s) failed.
            failed_kpis.append((kpi, str(_kpi_e)))
            value = unit = ground = None
        rows.append({"Company": inp.company, "Year": inp.year, "KPI_Name": kpi,
                     "Value": value, "Unit": unit, "GroundTruthCommonValue": ground})
    return rows, failed_kpis


def _upsert_longform(inp, out, units=None, corp_values=None,
                     country_values_gj=None, supp=None) -> str:
    """Overwrite this company+year's block in the long-form master, preserving
    canonical KPI ordering and every other company+year untouched. Returns a
    short status string (success or what went wrong) so the caller can show
    it to the user instead of the update failing invisibly."""
    order = _canonical_kpi_order()
    if not order:
        msg = "long-form skipped — canonical KPI order unavailable (couldn't read existing long-form master)"
        _log.warning("[long-form] %s", msg)
        return msg
    built_rows, failed_kpis = build_longform_rows(order, inp, out, units=units,
                                                  corp_values=corp_values,
                                                  country_values_gj=country_values_gj,
                                                  supp=supp)
    new_rows = pd.DataFrame(built_rows, columns=LONG_FORM_COLS)
    path = LONG_FORM_PATH
    if not path.exists():
        cands = [c for c in dl._get_csv_candidates()
                 if c.name.startswith("ESG_LONG_FORM_MASTER") and c.exists()]
        if cands:
            path = cands[0]
    path.parent.mkdir(parents=True, exist_ok=True)

    lock = FileLock(str(path.with_suffix(".lock")), timeout=cfg.FILELOCK_TIMEOUT)
    with lock:
        if path.exists():
            existing = pd.read_csv(path)
            keep = ~((existing["Company"] == inp.company)
                     & (existing["Year"].astype(str) == str(inp.year)))
            existing = existing[keep]
            combined = pd.concat([existing, new_rows], ignore_index=True)
        else:
            combined = new_rows.copy()
        # Stable block order (Company, Year) with canonical KPI order within.
        rank = {k: i for i, k in enumerate(order)}
        combined["_r"] = combined["KPI_Name"].map(rank).fillna(10**6).astype(int)
        combined = (combined.sort_values(["Company", "Year", "_r"], kind="stable")
                            .drop(columns="_r").reset_index(drop=True))
        combined.to_csv(path, index=False)

    if failed_kpis:
        _log.warning("[long-form] %d KPI(s) failed for %s %s: %s",
                     len(failed_kpis), inp.company, inp.year, failed_kpis)
        names = ", ".join(k for k, _ in failed_kpis[:5])
        more  = f" (+{len(failed_kpis)-5} more)" if len(failed_kpis) > 5 else ""
        return f"long-form updated, but {len(failed_kpis)} KPI(s) left blank: {names}{more}"
    return "long-form updated"